# -*- coding: utf-8 -*-
"""Dump sheet names and first rows of page_elements_with_defaults.xlsx to stdout."""
import json
import os
import sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("pip install openpyxl", file=sys.stderr)
    sys.exit(1)

base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
path = os.path.join(base, "Maanshan3DMap", "page_elements_with_defaults.xlsx")
if not os.path.isfile(path):
    path = os.path.join(base, "page_elements_with_defaults.xlsx")
if not os.path.isfile(path):
    print("File not found: page_elements_with_defaults.xlsx", file=sys.stderr)
    sys.exit(1)

wb = load_workbook(path, read_only=True)
out = {"path": path, "sheets": {}}
for name in wb.sheetnames:
    ws = wb[name]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        rows.append([str(c) if c is not None else "" for c in row])
        if i >= 20:
            break
    out["sheets"][name] = rows
print(json.dumps(out, ensure_ascii=False, indent=2))
